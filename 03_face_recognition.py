import cv2
import numpy as np
import os
import openpyxl


recognizer = cv2.face.LBPHFaceRecognizer_create()
recognizer.read('trainer/trainer.yml')


faceCascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

font = cv2.FONT_HERSHEY_SIMPLEX

id_to_name = {}
excel_file = 'user_data.xlsx'

if os.path.exists(excel_file):
    workbook = openpyxl.load_workbook(excel_file)
    if 'User Data' in workbook.sheetnames:
        sheet = workbook['User Data']
        for row in sheet.iter_rows(min_row=2, values_only=True):  
            try:
                face_id = int(row[0])
                username = str(row[1]).strip()
                id_to_name[face_id] = username
                print(f"[Excel] Loaded: ID={face_id}, Username='{username}'")
            except Exception as e:
                print(f"[Excel] Skipping row due to error: {e}")
    else:
        print("[ERROR] 'User Data' sheet not found in Excel.")
else:
    print("[ERROR] Excel file 'user_data.xlsx' not found.")

cam = cv2.VideoCapture(0)
cam.set(3, 640)  
cam.set(4, 480) 

minW = 0.1 * cam.get(3)
minH = 0.1 * cam.get(4)

print("\n[INFO] Starting camera. Press ESC to quit.")

while True:
    ret, img = cam.read()
    if not ret:
        print("[ERROR] Failed to grab frame.")
        break

    img = cv2.flip(img, 1)  
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    faces = faceCascade.detectMultiScale(
        gray,
        scaleFactor=1.2,
        minNeighbors=5,
        minSize=(int(minW), int(minH))
    )

    for (x, y, w, h) in faces:
        cv2.rectangle(img, (x, y), (x + w, y + h), (0, 255, 0), 2)

        id, confidence = recognizer.predict(gray[y:y + h, x:x + w])
        print(f"[Prediction] ID={id}, Confidence={confidence:.2f}")

        
        if confidence < 70: 
            name = id_to_name.get(id, "Unknown")
        else:
            name = "Unknown"

        print(f"[Lookup] id_to_name[{id}] = {name}")

        confidence_text = f"{round(100 - confidence)}%"

       
        cv2.putText(img, str(name), (x + 5, y - 5), font, 1, (255, 255, 255), 2)
        cv2.putText(img, confidence_text, (x + 5, y + h - 5), font, 1, (255, 255, 0), 1)

    cv2.imshow('camera', img)

    k = cv2.waitKey(10) & 0xff
    if k == 27: 
        break

print("\n[INFO] Exiting Program and cleaning up.")
cam.release()
cv2.destroyAllWindows()
