import cv2
import os
import openpyxl
from datetime import datetime

cam = cv2.VideoCapture(0)
cam.set(3, 640)  
cam.set(4, 480) 

face_detector = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')


face_id = input('\nEnter numeric user ID: ')
username = input('Enter username: ')
email = input('Enter email address: ')

print("\n[INFO] Initializing face capture. Look at the camera...")


if not os.path.exists('dataset'):
    os.makedirs('dataset')


excel_file = 'user_data.xlsx'
if not os.path.exists(excel_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "User Data"
    sheet.append(["Face ID", "Username", "Email", "Timestamp", "Image Count"])
    workbook.save(excel_file)


workbook = openpyxl.load_workbook(excel_file)
if "User Data" in workbook.sheetnames:
    sheet = workbook["User Data"]
else:
    sheet = workbook.create_sheet("User Data")
    sheet.append(["Face ID", "Username", "Email", "Timestamp", "Image Count"])
    workbook.save(excel_file)


if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
    std = workbook["Sheet"]
    workbook.remove(std)

count = 0
while True:
    ret, img = cam.read()
    img = cv2.flip(img, 1)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    faces = face_detector.detectMultiScale(gray, 1.3, 5)

    for (x, y, w, h) in faces:
        cv2.rectangle(img, (x, y), (x + w, y + h), (255, 0, 0), 2)
        count += 1

        img_name = f"dataset/User.{face_id}.{count}.jpg"
        cv2.imwrite(img_name, gray[y:y + h, x:x + w])

    cv2.imshow('image', img)

    k = cv2.waitKey(100) & 0xff
    if k == 27:  
        break
    elif count >= 30:  
        break

timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
sheet.append([face_id, username, email, timestamp, count])
workbook.save(excel_file)

print("\n[INFO] Exiting program and cleaning up.")
cam.release()
cv2.destroyAllWindows()
